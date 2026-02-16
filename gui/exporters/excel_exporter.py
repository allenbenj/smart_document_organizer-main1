"""
Enhanced Excel Exporter Module

Provides comprehensive Excel export functionality with multiple sheets,
formatting, charts, and advanced features.

Features:
- Multi-sheet workbooks (Entities, Documents, Statistics, Proposals)
- Professional formatting with colors, fonts, borders
- Auto-column width and freeze panes
- Summary sheet with formulas and aggregations
- Conditional formatting for confidence scores
- Charts and visualizations embedded in Excel
- Template-based exports
"""

import os
from typing import List, Dict, Optional, Any
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None


class ExcelExporter:
    """
    Enhanced Excel exporter with multi-sheet support and formatting.
    
    Creates professional Excel workbooks with multiple sheets, formatting,
    and embedded charts.
    """
    
    # Color scheme
    HEADER_COLOR = "2C3E50"  # Dark blue
    SUBHEADER_COLOR = "34495E"  # Medium blue
    ALT_ROW_COLOR = "ECF0F1"  # Light gray
    CONFIDENCE_HIGH = "2ECC71"  # Green
    CONFIDENCE_MEDIUM = "F39C12"  # Orange
    CONFIDENCE_LOW = "E74C3C"  # Red
    
    def __init__(self):
        """Initialize the Excel exporter."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl not installed. Install with: pip install openpyxl")
        
        self.workbook: Optional[Workbook] = None
        self.filename: Optional[str] = None
        
    def create_workbook(self, filename: str) -> Workbook:
        """
        Create a new Excel workbook.
        
        Args:
            filename: Path to save the workbook
            
        Returns:
            The created workbook
        """
        self.filename = filename
        self.workbook = Workbook()
        
        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            del self.workbook["Sheet"]
        
        return self.workbook
        
    def export_full_report(
        self,
        filename: str,
        entities: List[Dict],
        documents: List[Dict],
        proposals: List[Dict],
        statistics: Dict,
        memory_stats: Optional[Dict] = None
    ) -> str:
        """
        Export a complete report with multiple sheets.
        
        Args:
            filename: Output file path
            entities: List of extracted entities
            documents: List of processed documents
            proposals: List of entity proposals
            statistics: Processing statistics
            memory_stats: Memory system statistics (optional)
            
        Returns:
            Path to the created file
        """
        self.create_workbook(filename)
        
        # Create all sheets
        self._create_summary_sheet(entities, documents, proposals, statistics)
        self._create_entities_sheet(entities)
        self._create_documents_sheet(documents)
        self._create_proposals_sheet(proposals)
        self._create_statistics_sheet(statistics, memory_stats)
        
        # Save workbook
        self.workbook.save(filename)
        return filename
        
    def _create_summary_sheet(
        self,
        entities: List[Dict],
        documents: List[Dict],
        proposals: List[Dict],
        statistics: Dict
    ):
        """Create the summary overview sheet."""
        ws = self.workbook.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "Document Organization Report"
        ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.HEADER_COLOR, 
                                     end_color=self.HEADER_COLOR, 
                                     fill_type="solid")
        ws.merge_cells('A1:D1')
        
        # Metadata
        row = 3
        ws[f'A{row}'] = "Report Date:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Total Documents:"
        ws[f'B{row}'] = len(documents)
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Total Entities:"
        ws[f'B{row}'] = len(entities)
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Pending Proposals:"
        ws[f'B{row}'] = len([p for p in proposals if p.get('status') == 'pending'])
        ws[f'A{row}'].font = Font(bold=True)
        
        # Entity Type Summary
        row += 3
        ws[f'A{row}'] = "Entity Types Summary"
        ws[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color=self.SUBHEADER_COLOR,
                                         end_color=self.SUBHEADER_COLOR,
                                         fill_type="solid")
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        headers = ["Entity Type", "Count", "Percentage"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=self.ALT_ROW_COLOR,
                                   end_color=self.ALT_ROW_COLOR,
                                   fill_type="solid")
        
        # Count entities by type
        type_counts = {}
        for entity in entities:
            entity_type = entity.get('type', 'UNKNOWN')
            type_counts[entity_type] = type_counts.get(entity_type, 0) + 1
        
        total_entities = len(entities) or 1
        for entity_type, count in sorted(type_counts.items(), key=lambda x: x[1], reverse=True):
            row += 1
            ws[f'A{row}'] = entity_type
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"=B{row}/{total_entities}"
            ws[f'C{row}'].number_format = '0.0%'
        
        # Add chart
        if len(type_counts) > 0:
            chart = PieChart()
            chart.title = "Entity Distribution"
            chart.height = 10
            chart.width = 15
            
            data_start = row - len(type_counts) + 1
            data = Reference(ws, min_col=2, min_row=data_start, max_row=row)
            labels = Reference(ws, min_col=1, min_row=data_start, max_row=row)
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(labels)
            ws.add_chart(chart, f'E{data_start}')
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
        
    def _create_entities_sheet(self, entities: List[Dict]):
        """Create the entities detail sheet."""
        ws = self.workbook.create_sheet("Entities")
        
        # Headers
        headers = ["ID", "Text", "Type", "Confidence", "Source Document", "Context", "Metadata"]
        self._write_header_row(ws, 1, headers)
        
        # Data rows
        for row_idx, entity in enumerate(entities, 2):
            ws[f'A{row_idx}'] = entity.get('id', '')
            ws[f'B{row_idx}'] = entity.get('text', '')
            ws[f'C{row_idx}'] = entity.get('type', '')
            
            confidence = entity.get('confidence', 0.0)
            ws[f'D{row_idx}'] = confidence
            ws[f'D{row_idx}'].number_format = '0.0%'
            
            # Conditional formatting for confidence
            if confidence >= 0.8:
                fill_color = self.CONFIDENCE_HIGH
            elif confidence >= 0.5:
                fill_color = self.CONFIDENCE_MEDIUM
            else:
                fill_color = self.CONFIDENCE_LOW
            
            ws[f'D{row_idx}'].fill = PatternFill(start_color=fill_color,
                                                  end_color=fill_color,
                                                  fill_type="solid")
            
            ws[f'E{row_idx}'] = entity.get('source', '')
            ws[f'F{row_idx}'] = entity.get('context', '')[:100]  # Truncate context
            ws[f'G{row_idx}'] = str(entity.get('metadata', {}))
            
            # Alternate row colors
            if row_idx % 2 == 0:
                for col in range(1, 8):
                    if col != 4:  # Skip confidence column (already colored)
                        ws.cell(row_idx, col).fill = PatternFill(
                            start_color=self.ALT_ROW_COLOR,
                            end_color=self.ALT_ROW_COLOR,
                            fill_type="solid"
                        )
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add table
        if len(entities) > 0:
            table = Table(displayName="EntitiesTable", ref=f"A1:G{len(entities) + 1}")
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
        
    def _create_documents_sheet(self, documents: List[Dict]):
        """Create the documents detail sheet."""
        ws = self.workbook.create_sheet("Documents")
        
        # Headers
        headers = ["File Name", "Path", "Type", "Size (KB)", "Entities Found", 
                  "Processing Time (s)", "Status", "Date Processed"]
        self._write_header_row(ws, 1, headers)
        
        # Data rows
        for row_idx, doc in enumerate(documents, 2):
            ws[f'A{row_idx}'] = doc.get('name', '')
            ws[f'B{row_idx}'] = doc.get('path', '')
            ws[f'C{row_idx}'] = doc.get('type', '')
            
            # File size
            size_bytes = doc.get('size', 0)
            ws[f'D{row_idx}'] = size_bytes / 1024  # Convert to KB
            ws[f'D{row_idx}'].number_format = '0.00'
            
            ws[f'E{row_idx}'] = doc.get('entity_count', 0)
            
            # Processing time
            proc_time = doc.get('processing_time', 0)
            ws[f'F{row_idx}'] = proc_time
            ws[f'F{row_idx}'].number_format = '0.000'
            
            ws[f'G{row_idx}'] = doc.get('status', 'Unknown')
            
            # Date processed
            date_str = doc.get('date_processed', '')
            if date_str:
                ws[f'H{row_idx}'] = date_str
            
            # Alternate row colors
            if row_idx % 2 == 0:
                for col in range(1, 9):
                    ws.cell(row_idx, col).fill = PatternFill(
                        start_color=self.ALT_ROW_COLOR,
                        end_color=self.ALT_ROW_COLOR,
                        fill_type="solid"
                    )
        
        # Freeze header
        ws.freeze_panes = 'A2'
        
        # Add table
        if len(documents) > 0:
            table = Table(displayName="DocumentsTable", ref=f"A1:H{len(documents) + 1}")
            style = TableStyleInfo(
                name="TableStyleMedium3",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
        
    def _create_proposals_sheet(self, proposals: List[Dict]):
        """Create the entity proposals sheet."""
        ws = self.workbook.create_sheet("Proposals")
        
        # Headers
        headers = ["ID", "Entity Text", "Type", "Confidence", "Status", 
                  "Source", "Context", "Reviewer", "Review Date"]
        self._write_header_row(ws, 1, headers)
        
        # Data rows
        for row_idx, proposal in enumerate(proposals, 2):
            ws[f'A{row_idx}'] = proposal.get('id', '')
            ws[f'B{row_idx}'] = proposal.get('text', '')
            ws[f'C{row_idx}'] = proposal.get('type', '')
            
            confidence = proposal.get('confidence', 0.0)
            ws[f'D{row_idx}'] = confidence
            ws[f'D{row_idx}'].number_format = '0.0%'
            
            # Conditional formatting for confidence
            if confidence >= 0.8:
                fill_color = self.CONFIDENCE_HIGH
            elif confidence >= 0.5:
                fill_color = self.CONFIDENCE_MEDIUM
            else:
                fill_color = self.CONFIDENCE_LOW
            
            ws[f'D{row_idx}'].fill = PatternFill(start_color=fill_color,
                                                  end_color=fill_color,
                                                  fill_type="solid")
            
            status = proposal.get('status', 'pending')
            ws[f'E{row_idx}'] = status.upper()
            
            # Status color coding
            status_colors = {
                'approved': self.CONFIDENCE_HIGH,
                'rejected': self.CONFIDENCE_LOW,
                'pending': self.CONFIDENCE_MEDIUM
            }
            ws[f'E{row_idx}'].fill = PatternFill(
                start_color=status_colors.get(status, 'FFFFFF'),
                end_color=status_colors.get(status, 'FFFFFF'),
                fill_type="solid"
            )
            
            ws[f'F{row_idx}'] = proposal.get('source', '')
            ws[f'G{row_idx}'] = proposal.get('context', '')[:100]
            ws[f'H{row_idx}'] = proposal.get('reviewer', '')
            ws[f'I{row_idx}'] = proposal.get('review_date', '')
            
            # Alternate row colors (except confidence and status)
            if row_idx % 2 == 0:
                for col in [1, 2, 3, 6, 7, 8, 9]:
                    ws.cell(row_idx, col).fill = PatternFill(
                        start_color=self.ALT_ROW_COLOR,
                        end_color=self.ALT_ROW_COLOR,
                        fill_type="solid"
                    )
        
        # Freeze header
        ws.freeze_panes = 'A2'
        
        # Add table
        if len(proposals) > 0:
            table = Table(displayName="ProposalsTable", ref=f"A1:I{len(proposals) + 1}")
            style = TableStyleInfo(
                name="TableStyleMedium4",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
        
    def _create_statistics_sheet(self, statistics: Dict, memory_stats: Optional[Dict] = None):
        """Create the statistics sheet."""
        ws = self.workbook.create_sheet("Statistics")
        
        # Processing Statistics Section
        row = 1
        ws[f'A{row}'] = "Processing Statistics"
        ws[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color=self.HEADER_COLOR,
                                         end_color=self.HEADER_COLOR,
                                         fill_type="solid")
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        stats_data = [
            ("Total Documents Processed", statistics.get('total_documents', 0)),
            ("Total Entities Extracted", statistics.get('total_entities', 0)),
            ("Unique Entity Types", statistics.get('unique_types', 0)),
            ("Average Confidence", statistics.get('avg_confidence', 0.0)),
            ("Average Processing Time (s)", statistics.get('avg_processing_time', 0.0)),
            ("Success Rate", statistics.get('success_rate', 0.0)),
        ]
        
        for label, value in stats_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            
            # Format percentages and decimals
            if isinstance(value, float):
                if label.endswith('Rate'):
                    ws[f'B{row}'].number_format = '0.0%'
                else:
                    ws[f'B{row}'].number_format = '0.000'
            
            row += 1
        
        # Memory Statistics Section (if provided)
        if memory_stats:
            row += 2
            ws[f'A{row}'] = "Memory System Statistics"
            ws[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color=self.SUBHEADER_COLOR,
                                             end_color=self.SUBHEADER_COLOR,
                                             fill_type="solid")
            ws.merge_cells(f'A{row}:B{row}')
            
            row += 1
            mem_data = [
                ("Total Records", memory_stats.get('total_records', 0)),
                ("Vector Backend", memory_stats.get('backend', 'Unknown')),
                ("Cache Hits", memory_stats.get('cache_hits', 0)),
                ("Cache Misses", memory_stats.get('cache_misses', 0)),
                ("Cache Hit Rate", memory_stats.get('cache_hit_rate', 0.0)),
            ]
            
            for label, value in mem_data:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                
                if isinstance(value, float) and label.endswith('Rate'):
                    ws[f'B{row}'].number_format = '0.0%'
                
                row += 1
        
        # Entity Type Distribution Chart
        if 'entity_type_distribution' in statistics:
            row += 2
            ws[f'A{row}'] = "Entity Type Distribution"
            ws[f'A{row}'].font = Font(size=12, bold=True)
            ws.merge_cells(f'A{row}:C{row}')
            
            row += 1
            ws[f'A{row}'] = "Type"
            ws[f'B{row}'] = "Count"
            ws[f'C{row}'] = "Percentage"
            for col in range(1, 4):
                ws.cell(row, col).font = Font(bold=True)
                ws.cell(row, col).fill = PatternFill(start_color=self.ALT_ROW_COLOR,
                                                     end_color=self.ALT_ROW_COLOR,
                                                     fill_type="solid")
            
            chart_start = row
            for entity_type, count in statistics['entity_type_distribution'].items():
                row += 1
                ws[f'A{row}'] = entity_type
                ws[f'B{row}'] = count
                ws[f'C{row}'] = f"=B{row}/SUM(B{chart_start+1}:B{row})"
                ws[f'C{row}'].number_format = '0.0%'
            
            # Add bar chart
            chart = BarChart()
            chart.title = "Entity Type Distribution"
            chart.x_axis.title = "Entity Type"
            chart.y_axis.title = "Count"
            
            data = Reference(ws, min_col=2, min_row=chart_start, max_row=row)
            cats = Reference(ws, min_col=1, min_row=chart_start+1, max_row=row)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 20
            
            ws.add_chart(chart, f'E{chart_start}')
        
        # Auto-adjust columns
        self._auto_adjust_columns(ws)
        
    def _write_header_row(self, ws, row: int, headers: List[str]):
        """Write a formatted header row."""
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row, col_idx, header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color=self.HEADER_COLOR,
                                   end_color=self.HEADER_COLOR,
                                   fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
            
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width
            
    def export_entities_only(self, filename: str, entities: List[Dict]) -> str:
        """
        Export only entities to Excel.
        
        Args:
            filename: Output file path
            entities: List of entities
            
        Returns:
            Path to the created file
        """
        self.create_workbook(filename)
        self._create_entities_sheet(entities)
        self.workbook.save(filename)
        return filename
        
    def export_proposals_only(self, filename: str, proposals: List[Dict]) -> str:
        """
        Export only proposals to Excel.
        
        Args:
            filename: Output file path
            proposals: List of proposals
            
        Returns:
            Path to the created file
        """
        self.create_workbook(filename)
        self._create_proposals_sheet(proposals)
        self.workbook.save(filename)
        return filename


# Example usage
if __name__ == "__main__":
    # Sample data
    entities = [
        {
            "id": "e1",
            "text": "John Doe",
            "type": "PERSON",
            "confidence": 0.95,
            "source": "document1.pdf",
            "context": "John Doe signed the contract",
            "metadata": {"page": 1}
        },
        {
            "id": "e2",
            "text": "Acme Corp",
            "type": "ORG",
            "confidence": 0.88,
            "source": "document1.pdf",
            "context": "Acme Corp is headquartered in",
            "metadata": {"page": 2}
        },
    ]
    
    documents = [
        {
            "name": "document1.pdf",
            "path": "/docs/document1.pdf",
            "type": "PDF",
            "size": 102400,
            "entity_count": 15,
            "processing_time": 2.5,
            "status": "Completed",
            "date_processed": "2024-01-15"
        }
    ]
    
    proposals = [
        {
            "id": "p1",
            "text": "Jane Smith",
            "type": "PERSON",
            "confidence": 0.75,
            "status": "pending",
            "source": "document2.pdf",
            "context": "Jane Smith reviewed the proposal"
        }
    ]
    
    statistics = {
        "total_documents": 25,
        "total_entities": 150,
        "unique_types": 8,
        "avg_confidence": 0.85,
        "avg_processing_time": 1.8,
        "success_rate": 0.96,
        "entity_type_distribution": {
            "PERSON": 45,
            "ORG": 32,
            "GPE": 28,
            "DATE": 25,
            "MONEY": 20
        }
    }
    
    memory_stats = {
        "total_records": 90,
        "backend": "chromadb",
        "cache_hits": 45,
        "cache_misses": 12,
        "cache_hit_rate": 0.79
    }
    
    # Export full report
    exporter = ExcelExporter()
    output_file = "sample_report.xlsx"
    exporter.export_full_report(
        output_file,
        entities,
        documents,
        proposals,
        statistics,
        memory_stats
    )
    
    print(f"Report exported to: {output_file}")
